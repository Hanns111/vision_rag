"""
Generador del Excel de validación humana.

- Entrada: lista de `DocumentoValidacion` (una por archivo) + lista de `ExpedienteValidacion`.
- Salida: `data/piloto_ocr/metrics/validacion_expedientes.xlsx`, 3 hojas:
    * documentos  → una fila por archivo (columnas de sistema + columnas humanas vacías)
    * expedientes → resumen por expediente_id
    * errores     → subconjunto de documentos con estado_procesamiento = error
- Upsert por (expediente_id, archivo): si la fila ya existe, actualiza solo las
  columnas de sistema y preserva lo que el humano haya escrito.

Columnas de sistema: se REESCRIBEN en cada export.
Columnas humanas (tipo_correcto, monto_correcto, fecha_correcta, ruc_correcto,
observaciones, validacion_final): nunca se sobrescriben si ya tienen valor.
"""

from __future__ import annotations

from dataclasses import dataclass, field
from pathlib import Path
from typing import Any


@dataclass
class DocumentoValidacion:
    expediente_id: str
    archivo: str
    ruta_origen: str
    tipo_documento_detectado: str
    confianza_tipo: float
    monto_detectado: str | None
    fecha_detectada: str | None
    ruc_detectado: str | None
    razon_social_detectada: str | None
    numero_documento_detectado: str | None
    tipo_gasto_detectado: str | None
    texto_extraido_resumen: str
    estado_procesamiento: str  # "ok" | "bajo_confianza" | "error"
    nota_sistema: str = ""     # mensajes internos (overrides, errores, etc.)
    # --- Validación normativa: firmas Anexo 3 (opcional; N/A si no es rendición) ---
    validacion_firmas: str = ""       # "firmas_anexo3" | "" (no aplica)
    estado_firmas: str = ""           # CONFORME | OBSERVADO | INSUFICIENTE_EVIDENCIA | ""
    errores_firmas: str = ""          # lista separada por "; "
    confianza_firmas: float | str = ""
    # --- Resolución de identidad administrativa (SINAD, SIAF, AÑO, EXP) ---
    expediente_detectado: str = ""        # id_canonico ganador (ej. "SINAD-250235")
    sinad_detectado: str = ""             # valor sin prefijo (ej. "250235")
    siaf_detectado: str = ""
    anio_detectado: str = ""
    confianza_expediente: float | str = ""
    confianza_sinad: float | str = ""
    confianza_siaf: float | str = ""
    conflicto_expediente: str = ""        # "Sí" | "No" | ""
    observaciones_expediente: str = ""    # lista unida con "; "


@dataclass
class CandidatoResolucion:
    """Fila de la hoja `resolucion_ids` (una por candidato por expediente)."""
    expediente_carpeta: str
    id_canonico: str
    tipo: str
    frecuencia: int
    score_total: float
    coincide_con_carpeta: str             # "Sí" | "No"
    es_ganador: str                       # "Sí" | "No"
    estado_resolucion: str                # estado global del expediente
    fuentes: str                          # resumen de archivos + páginas


@dataclass
class ComprobanteExcel:
    """Fila de la hoja `comprobantes` (una por comprobante por expediente)."""
    expediente_id: str
    archivo: str
    pagina_inicio: int
    pagina_fin: int
    tipo: str
    ruc: str
    razon_social: str
    serie_numero: str
    fecha: str
    monto_total: str
    moneda: str
    monto_igv: str
    confianza: float | str
    texto_resumen: str
    # columnas humanas
    monto_correcto: str = ""
    ruc_correcto: str = ""
    proveedor_correcto: str = ""
    observaciones: str = ""
    validacion_final: str = ""


@dataclass
class ExpedienteValidacion:
    expediente_id: str
    ruta_origen: str
    ruta_destino: str
    n_documentos: int
    n_ok: int
    n_bajo_confianza: int
    n_error: int
    tipos_detectados: str       # lista separada por coma
    fecha_ingesta: str


_COLS_SISTEMA_DOC = [
    "expediente_id",
    "archivo",
    "ruta_origen",
    "tipo_documento_detectado",
    "confianza_tipo",
    "monto_detectado",
    "fecha_detectada",
    "ruc_detectado",
    "razon_social_detectada",
    "numero_documento_detectado",
    "tipo_gasto_detectado",
    "texto_extraido_resumen",
    "estado_procesamiento",
    "nota_sistema",
    # validación normativa firmas Anexo 3
    "validacion_firmas",
    "estado_firmas",
    "errores_firmas",
    "confianza_firmas",
    # resolución de identidad (SINAD, SIAF, AÑO, EXP)
    "expediente_detectado",
    "sinad_detectado",
    "siaf_detectado",
    "anio_detectado",
    "confianza_expediente",
    "confianza_sinad",
    "confianza_siaf",
    "conflicto_expediente",
    "observaciones_expediente",
]

_COLS_RESOLUCION_IDS = [
    "expediente_carpeta",
    "id_canonico",
    "tipo",
    "frecuencia",
    "score_total",
    "coincide_con_carpeta",
    "es_ganador",
    "estado_resolucion",
    "fuentes",
]

_COLS_SISTEMA_COMP = [
    "expediente_id",
    "archivo",
    "pagina_inicio",
    "pagina_fin",
    "tipo",
    "ruc",
    "razon_social",
    "serie_numero",
    "fecha",
    "monto_total",
    "moneda",
    "monto_igv",
    "confianza",
    "texto_resumen",
]
_COLS_HUMANAS_COMP = [
    "monto_correcto",
    "ruc_correcto",
    "proveedor_correcto",
    "observaciones",
    "validacion_final",
]
_COLS_COMP = _COLS_SISTEMA_COMP + _COLS_HUMANAS_COMP
_COLS_HUMANAS_DOC = [
    "tipo_correcto",
    "monto_correcto",
    "fecha_correcta",
    "ruc_correcto",
    "observaciones",
    "validacion_final",
]
_COLS_DOC = _COLS_SISTEMA_DOC + _COLS_HUMANAS_DOC

_COLS_EXPEDIENTES = [
    "expediente_id",
    "ruta_origen",
    "ruta_destino",
    "n_documentos",
    "n_ok",
    "n_bajo_confianza",
    "n_error",
    "tipos_detectados",
    "fecha_ingesta",
]


def _asdict_doc(d: DocumentoValidacion) -> dict[str, Any]:
    return {
        "expediente_id": d.expediente_id,
        "archivo": d.archivo,
        "ruta_origen": d.ruta_origen,
        "tipo_documento_detectado": d.tipo_documento_detectado,
        "confianza_tipo": d.confianza_tipo,
        "monto_detectado": d.monto_detectado,
        "fecha_detectada": d.fecha_detectada,
        "ruc_detectado": d.ruc_detectado,
        "razon_social_detectada": d.razon_social_detectada,
        "numero_documento_detectado": d.numero_documento_detectado,
        "tipo_gasto_detectado": d.tipo_gasto_detectado,
        "texto_extraido_resumen": d.texto_extraido_resumen,
        "estado_procesamiento": d.estado_procesamiento,
        "nota_sistema": d.nota_sistema,
        "validacion_firmas": d.validacion_firmas,
        "estado_firmas": d.estado_firmas,
        "errores_firmas": d.errores_firmas,
        "confianza_firmas": d.confianza_firmas,
        "expediente_detectado": d.expediente_detectado,
        "sinad_detectado": d.sinad_detectado,
        "siaf_detectado": d.siaf_detectado,
        "anio_detectado": d.anio_detectado,
        "confianza_expediente": d.confianza_expediente,
        "confianza_sinad": d.confianza_sinad,
        "confianza_siaf": d.confianza_siaf,
        "conflicto_expediente": d.conflicto_expediente,
        "observaciones_expediente": d.observaciones_expediente,
    }


def _asdict_cand(c: CandidatoResolucion) -> dict[str, Any]:
    return {
        "expediente_carpeta": c.expediente_carpeta,
        "id_canonico": c.id_canonico,
        "tipo": c.tipo,
        "frecuencia": c.frecuencia,
        "score_total": c.score_total,
        "coincide_con_carpeta": c.coincide_con_carpeta,
        "es_ganador": c.es_ganador,
        "estado_resolucion": c.estado_resolucion,
        "fuentes": c.fuentes,
    }


def _asdict_comp(c: ComprobanteExcel) -> dict[str, Any]:
    return {
        "expediente_id": c.expediente_id,
        "archivo": c.archivo,
        "pagina_inicio": c.pagina_inicio,
        "pagina_fin": c.pagina_fin,
        "tipo": c.tipo,
        "ruc": c.ruc,
        "razon_social": c.razon_social,
        "serie_numero": c.serie_numero,
        "fecha": c.fecha,
        "monto_total": c.monto_total,
        "moneda": c.moneda,
        "monto_igv": c.monto_igv,
        "confianza": c.confianza,
        "texto_resumen": c.texto_resumen,
        "monto_correcto": c.monto_correcto,
        "ruc_correcto": c.ruc_correcto,
        "proveedor_correcto": c.proveedor_correcto,
        "observaciones": c.observaciones,
        "validacion_final": c.validacion_final,
    }


def _asdict_exp(e: ExpedienteValidacion) -> dict[str, Any]:
    return {
        "expediente_id": e.expediente_id,
        "ruta_origen": e.ruta_origen,
        "ruta_destino": e.ruta_destino,
        "n_documentos": e.n_documentos,
        "n_ok": e.n_ok,
        "n_bajo_confianza": e.n_bajo_confianza,
        "n_error": e.n_error,
        "tipos_detectados": e.tipos_detectados,
        "fecha_ingesta": e.fecha_ingesta,
    }


def _cargar_filas_existentes(
    path: Path, sheet: str, cols: list[str], key_cols: tuple[str, ...]
) -> dict[tuple, dict[str, Any]]:
    """Lee el xlsx si existe y devuelve {clave: fila_completa}. Clave = tupla de key_cols."""
    if not path.exists():
        return {}
    from openpyxl import load_workbook

    wb = load_workbook(path)
    if sheet not in wb.sheetnames:
        return {}
    ws = wb[sheet]
    headers = [c.value for c in ws[1]] if ws.max_row >= 1 else []
    if not headers:
        return {}
    out: dict[tuple, dict[str, Any]] = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        rec = {h: v for h, v in zip(headers, row) if h}
        if not rec:
            continue
        key = tuple(rec.get(k) for k in key_cols)
        if all(k is not None for k in key):
            out[key] = rec
    return out


def _escribir_hoja(
    wb: Any, sheet: str, cols: list[str], filas: list[dict[str, Any]]
) -> None:
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    if sheet in wb.sheetnames:
        del wb[sheet]
    ws = wb.create_sheet(sheet)
    ws.append(cols)

    header_fill = PatternFill("solid", fgColor="305496")
    header_font = Font(bold=True, color="FFFFFF")
    human_fill = PatternFill("solid", fgColor="FFF2CC")  # amarillo suave

    for i, col in enumerate(cols, start=1):
        c = ws.cell(row=1, column=i)
        c.fill = header_fill
        c.font = header_font
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # filas
    for row_idx, f in enumerate(filas, start=2):
        for i, col in enumerate(cols, start=1):
            v = f.get(col)
            c = ws.cell(row=row_idx, column=i, value=v)
            if col in _COLS_HUMANAS_DOC or col in _COLS_HUMANAS_COMP:
                c.fill = human_fill
            if col in ("texto_extraido_resumen", "texto_resumen"):
                c.alignment = Alignment(wrap_text=True, vertical="top")

    # anchos de columna aproximados
    anchos = {
        "expediente_carpeta": 26,
        "id_canonico": 22,
        "tipo": 12,
        "frecuencia": 10,
        "score_total": 12,
        "coincide_con_carpeta": 18,
        "es_ganador": 10,
        "estado_resolucion": 22,
        "fuentes": 60,
        "expediente_id": 26,
        "archivo": 50,
        "ruta_origen": 60,
        "tipo_documento_detectado": 20,
        "confianza_tipo": 14,
        "monto_detectado": 16,
        "fecha_detectada": 14,
        "ruc_detectado": 16,
        "razon_social_detectada": 30,
        "numero_documento_detectado": 22,
        "tipo_gasto_detectado": 20,
        "texto_extraido_resumen": 60,
        "estado_procesamiento": 18,
        "nota_sistema": 40,
        "validacion_firmas": 18,
        "estado_firmas": 22,
        "errores_firmas": 40,
        "confianza_firmas": 16,
        "expediente_detectado": 22,
        "sinad_detectado": 16,
        "siaf_detectado": 16,
        "anio_detectado": 12,
        "confianza_expediente": 18,
        "confianza_sinad": 16,
        "confianza_siaf": 16,
        "conflicto_expediente": 18,
        "observaciones_expediente": 40,
        "pagina_inicio": 10,
        "pagina_fin": 10,
        "ruc": 14,
        "razon_social": 32,
        "serie_numero": 16,
        "monto_total": 14,
        "moneda": 8,
        "monto_igv": 12,
        "confianza": 12,
        "monto_correcto": 14,
        "ruc_correcto": 14,
        "proveedor_correcto": 28,
        "tipo_correcto": 18,
        "monto_correcto": 14,
        "fecha_correcta": 14,
        "ruc_correcto": 16,
        "observaciones": 40,
        "validacion_final": 18,
        "ruta_destino": 60,
        "n_documentos": 14,
        "n_ok": 10,
        "n_bajo_confianza": 16,
        "n_error": 10,
        "tipos_detectados": 30,
        "fecha_ingesta": 24,
    }
    for i, col in enumerate(cols, start=1):
        ws.column_dimensions[get_column_letter(i)].width = anchos.get(col, 20)

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions


def exportar_excel(
    xlsx_path: Path | str,
    documentos: list[DocumentoValidacion],
    expedientes: list[ExpedienteValidacion],
    candidatos: list[CandidatoResolucion] | None = None,
    comprobantes: list[ComprobanteExcel] | None = None,
) -> Path:
    """Genera / actualiza el Excel preservando columnas humanas existentes."""
    from openpyxl import Workbook, load_workbook

    xlsx_path = Path(xlsx_path).resolve()
    xlsx_path.parent.mkdir(parents=True, exist_ok=True)

    key_cols = ("expediente_id", "archivo")
    prev = _cargar_filas_existentes(xlsx_path, "documentos", _COLS_DOC, key_cols)

    nuevas_docs: list[dict[str, Any]] = []
    claves_actualizadas: set[tuple] = set()
    for d in documentos:
        fila = _asdict_doc(d)
        # preservar columnas humanas si existían
        key = (d.expediente_id, d.archivo)
        if key in prev:
            for col in _COLS_HUMANAS_DOC:
                v = prev[key].get(col)
                if v not in (None, ""):
                    fila[col] = v
        for col in _COLS_HUMANAS_DOC:
            fila.setdefault(col, None)
        nuevas_docs.append(fila)
        claves_actualizadas.add(key)

    # conservar filas previas que NO se volvieron a procesar (otros expedientes)
    for key, rec in prev.items():
        if key in claves_actualizadas:
            continue
        fila = {col: rec.get(col) for col in _COLS_DOC}
        nuevas_docs.append(fila)

    # errores = subset
    filas_errores = [f for f in nuevas_docs if f.get("estado_procesamiento") == "error"]

    if xlsx_path.exists():
        wb = load_workbook(xlsx_path)
    else:
        wb = Workbook()
        # eliminar sheet por defecto que viene vacío
        if "Sheet" in wb.sheetnames:
            del wb["Sheet"]

    _escribir_hoja(wb, "documentos", _COLS_DOC, nuevas_docs)
    _escribir_hoja(wb, "expedientes", _COLS_EXPEDIENTES, [_asdict_exp(e) for e in expedientes])
    _escribir_hoja(wb, "errores", _COLS_DOC, filas_errores)
    if candidatos is not None:
        _escribir_hoja(
            wb,
            "resolucion_ids",
            _COLS_RESOLUCION_IDS,
            [_asdict_cand(c) for c in candidatos],
        )
    if comprobantes is not None:
        # Upsert: preservar columnas humanas previas por (expediente_id, archivo,
        # pagina_inicio, pagina_fin).
        prev_comp = _cargar_filas_existentes(
            xlsx_path,
            "comprobantes",
            _COLS_COMP,
            ("expediente_id", "archivo", "pagina_inicio", "pagina_fin"),
        )
        filas_comp: list[dict[str, Any]] = []
        for c in comprobantes:
            fila = _asdict_comp(c)
            key = (c.expediente_id, c.archivo, c.pagina_inicio, c.pagina_fin)
            if key in prev_comp:
                for col in _COLS_HUMANAS_COMP:
                    v = prev_comp[key].get(col)
                    if v not in (None, ""):
                        fila[col] = v
            for col in _COLS_HUMANAS_COMP:
                fila.setdefault(col, None)
            filas_comp.append(fila)
        _escribir_hoja(wb, "comprobantes", _COLS_COMP, filas_comp)

    orden = ["documentos", "expedientes", "errores", "resolucion_ids", "comprobantes"]
    wb._sheets = [wb[n] for n in orden if n in wb.sheetnames]
    wb.save(xlsx_path)
    return xlsx_path
