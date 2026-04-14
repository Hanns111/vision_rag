"""
PASO 3 — Mini experimento A/B (puntual, sin cambiar roadmap).

3 páginas piloto, motor tesseract vía pytesseract sobre raster 300 dpi.
A = gris sin CLAHE; B1 = gris + CLAHE; B2 = B1 + unsharp suave.

Salida: data/piloto_ocr/metrics/paso3_ab_linux_wsl/
"""

from __future__ import annotations

import csv
import json
import re
import sys
from datetime import date, datetime, timezone
from pathlib import Path
from typing import Any

_REPO = Path(__file__).resolve().parent.parent
_OUT = _REPO / "data" / "piloto_ocr" / "metrics" / "paso3_ab_linux_wsl"
_RAW = _REPO / "data" / "piloto_ocr" / "raw"
_LABELS = _REPO / "data" / "piloto_ocr" / "labels"

# (doc_id, nombre_pdf, page_1based)
PAGES: list[tuple[str, str, int]] = [
    ("sol-viat-debedsar-2026", "sol-viat-debedsar-2026.pdf", 2),
    ("rend-debedsar-amiquero-2026", "rend-debedsar-amiquero-2026.pdf", 5),
    ("rend-debedsar-amiquero-2026", "rend-debedsar-amiquero-2026.pdf", 19),
]

TARGET_FIELDS = ("serie_numero", "monto_subtotal", "monto_total", "monto_igv")
NON_TARGET_FIELDS = tuple(
    k
    for k in [
        "ruc_emisor",
        "tipo_documento",
        "fecha_emision",
        "moneda",
        "ruc_receptor",
        "razon_social_emisor",
    ]
)

CLAHE_CLIP = 2.0
CLAHE_TILE = (8, 8)
UNSHARP_SIGMA = 1.0
UNSHARP_AMOUNT = 0.6


def _page_bgr_300dpi(pdf_path: Path, page_1based: int):
    import cv2
    import fitz
    import numpy as np

    doc = fitz.open(str(pdf_path))
    page = doc.load_page(page_1based - 1)
    pix = page.get_pixmap(matrix=fitz.Matrix(300 / 72, 300 / 72), alpha=False)
    arr = np.frombuffer(pix.samples, dtype=np.uint8).reshape(pix.height, pix.width, pix.n)
    if pix.n == 3:
        arr = cv2.cvtColor(arr, cv2.COLOR_RGB2BGR)
    elif pix.n == 4:
        arr = cv2.cvtColor(arr, cv2.COLOR_RGBA2BGR)
    doc.close()
    return arr


def _gray_no_preprocess(bgr):
    import cv2

    return cv2.cvtColor(bgr, cv2.COLOR_BGR2GRAY)


def _preprocess_b1(gray):
    import cv2

    clahe = cv2.createCLAHE(clipLimit=CLAHE_CLIP, tileGridSize=CLAHE_TILE)
    return clahe.apply(gray)


def _preprocess_b2(gray_b1):
    import cv2

    blur = cv2.GaussianBlur(gray_b1, (0, 0), UNSHARP_SIGMA)
    return cv2.addWeighted(gray_b1, 1.0 + UNSHARP_AMOUNT, blur, -UNSHARP_AMOUNT, 0)


def _ocr_gray(gray) -> str:
    import pytesseract

    text = pytesseract.image_to_string(gray, lang="spa+eng")
    return (text or "").strip()


def _load_gold(doc_id: str, page_index: int) -> dict[str, Any]:
    p = _LABELS / f"{doc_id}_p{page_index}.json"
    with open(p, encoding="utf-8") as f:
        return json.load(f)


def _gold_evaluable(gold: Any) -> bool:
    if gold is None:
        return False
    if isinstance(gold, str) and not gold.strip():
        return False
    return True


def _paso4_candidate(gold: Any, pred_ok: bool, ocr_text: str) -> bool:
    if pred_ok or not _gold_evaluable(gold):
        return False
    g = str(gold).strip()
    if len(g) < 2:
        return False
    t = ocr_text or ""
    if g.upper() in t.upper():
        return True
    if re.sub(r"\s+", "", g).upper() in re.sub(r"\s+", "", t).upper():
        return True
    return False


def main() -> None:
    if hasattr(sys.stdout, "reconfigure"):
        try:
            sys.stdout.reconfigure(encoding="utf-8", errors="replace")
        except Exception:
            pass

    sys.path.insert(0, str(_REPO / "scripts"))
    from piloto_field_extract_minimal import extract_fields_minimal, gold_pred_match

    _OUT.mkdir(parents=True, exist_ok=True)
    (_OUT / "rasters").mkdir(exist_ok=True)
    (_OUT / "ocr_txt").mkdir(exist_ok=True)

    run_date = date.today().isoformat()
    ts = datetime.now(timezone.utc).astimezone().strftime("%Y-%m-%d %H:%M:%S %z")

    rows_detail: list[dict[str, Any]] = []
    texts_by_key: dict[tuple[str, int, str], str] = {}

    for doc_id, nombre_pdf, page_idx in PAGES:
        pdf_path = _RAW / nombre_pdf
        if not pdf_path.is_file():
            raise FileNotFoundError(pdf_path)

        bgr = _page_bgr_300dpi(pdf_path, page_idx)
        stem = f"{doc_id}_p{page_idx}"

        import cv2

        cv2.imwrite(str(_OUT / "rasters" / f"{stem}_A_bgr.png"), bgr)

        gray = _gray_no_preprocess(bgr)
        gray_b1 = _preprocess_b1(gray)
        gray_b2 = _preprocess_b2(gray_b1)

        cv2.imwrite(str(_OUT / "rasters" / f"{stem}_B1_clahe.png"), gray_b1)
        cv2.imwrite(str(_OUT / "rasters" / f"{stem}_B2_unsharp.png"), gray_b2)

        variants = {
            "A": gray,
            "B1": gray_b1,
            "B2": gray_b2,
        }

        gold_j = _load_gold(doc_id, page_idx)
        gold_campos = gold_j["campos"]

        for vname, gray_img in variants.items():
            txt = _ocr_gray(gray_img)
            texts_by_key[(doc_id, page_idx, vname)] = txt
            ptxt = _OUT / "ocr_txt" / f"{stem}_{vname}.txt"
            ptxt.write_text(txt, encoding="utf-8")

        preds = {v: extract_fields_minimal(texts_by_key[(doc_id, page_idx, v)]) for v in ("A", "B1", "B2")}
        text_a = texts_by_key[(doc_id, page_idx, "A")]

        all_fields = TARGET_FIELDS + NON_TARGET_FIELDS
        for campo in all_fields:
            gv = gold_campos.get(campo)
            ok_a = gold_pred_match(campo, gv, preds["A"].get(campo)) if _gold_evaluable(gv) else None
            ok_b1 = gold_pred_match(campo, gv, preds["B1"].get(campo)) if _gold_evaluable(gv) else None
            ok_b2 = gold_pred_match(campo, gv, preds["B2"].get(campo)) if _gold_evaluable(gv) else None

            def _delta(ok_b: bool | None, ok_a: bool | None) -> str:
                if ok_a is None or ok_b is None:
                    return "N/A"
                return str(int(ok_b) - int(ok_a))

            rows_detail.append(
                {
                    "archivo": nombre_pdf,
                    "doc_id": doc_id,
                    "page_index": page_idx,
                    "campo": campo,
                    "es_objetivo": "Sí" if campo in TARGET_FIELDS else "No",
                    "gold": gv if gv is not None else "",
                    "pred_A": preds["A"].get(campo) if preds["A"].get(campo) is not None else "",
                    "pred_B1": preds["B1"].get(campo) if preds["B1"].get(campo) is not None else "",
                    "pred_B2": preds["B2"].get(campo) if preds["B2"].get(campo) is not None else "",
                    "ok_A": "" if ok_a is None else ("1" if ok_a else "0"),
                    "ok_B1": "" if ok_b1 is None else ("1" if ok_b1 else "0"),
                    "ok_B2": "" if ok_b2 is None else ("1" if ok_b2 else "0"),
                    "delta_B1_vs_A": _delta(ok_b1, ok_a),
                    "delta_B2_vs_A": _delta(ok_b2, ok_a),
                    "paso4_candidato_A": "Sí"
                    if _gold_evaluable(gv)
                    and ok_a is False
                    and _paso4_candidate(gv, bool(ok_a), text_a)
                    else "No",
                }
            )

    # Métricas de decisión
    def _score_target(variant: str) -> tuple[int, int, int]:
        """(aciertos_target, net_vs_A_target, regresiones_no_target vs A)."""
        ok_key = f"ok_{variant}" if variant == "A" else f"ok_{variant}"  # noqa
        # Sum correct target fields where gold evaluable
        acc = 0
        net_b = 0
        reg_nt = 0
        for row in rows_detail:
            campo = row["campo"]
            ok_a = row["ok_A"]
            if ok_a == "":
                continue
            oa = ok_a == "1"
            if campo in TARGET_FIELDS:
                ob = row[f"ok_{variant}"] == "1" if row[f"ok_{variant}"] != "" else None
                if ob is not None:
                    if ob:
                        acc += 1
                    if variant != "A":
                        d = int(row[f"delta_{variant}_vs_A"])
                        net_b += d
            elif campo in NON_TARGET_FIELDS and variant != "A":
                ob = row[f"ok_{variant}"] == "1" if row[f"ok_{variant}"] != "" else None
                if ob is not None and oa and not ob:
                    reg_nt += 1
        return acc, net_b, reg_nt

    # Fix _score_target: for acc count per variant
    def target_correct_count(variant: str) -> int:
        k = "ok_A" if variant == "A" else f"ok_{variant}"
        n = 0
        for row in rows_detail:
            if row["campo"] not in TARGET_FIELDS:
                continue
            if row[k] == "":
                continue
            if row[k] == "1":
                n += 1
        return n

    def net_target_delta(variant: str) -> int:
        s = 0
        for row in rows_detail:
            if row["campo"] not in TARGET_FIELDS:
                continue
            d = row[f"delta_{variant}_vs_A"]
            if d == "N/A":
                continue
            s += int(d)
        return s

    def regressions_non_target(variant: str) -> int:
        r = 0
        for row in rows_detail:
            if row["campo"] not in NON_TARGET_FIELDS:
                continue
            if row["ok_A"] == "" or row[f"ok_{variant}"] == "":
                continue
            if row["ok_A"] == "1" and row[f"ok_{variant}"] == "0":
                r += 1
        return r

    a_tgt = target_correct_count("A")
    b1_net = net_target_delta("B1")
    b2_net = net_target_delta("B2")
    b1_reg = regressions_non_target("B1")
    b2_reg = regressions_non_target("B2")

    b1_ok = b1_net >= 3 and b1_reg <= 1
    b2_ok = b2_net >= 3 and b2_reg <= 1
    paso3_sigue = b1_ok or b2_ok
    mejor = "B1" if b1_ok else ("B2" if b2_ok else "—")

    csv_path = _OUT / "paso3_ab_delta_por_campo.csv"
    keys = list(rows_detail[0].keys()) if rows_detail else []
    with open(csv_path, "w", encoding="utf-8", newline="") as f:
        w = csv.DictWriter(f, fieldnames=keys)
        w.writeheader()
        w.writerows(rows_detail)

    # Formato largo: una fila por (página, campo, variante) — revisión humana
    rows_long: list[dict[str, Any]] = []
    for row in rows_detail:
        base = {
            "archivo": row["archivo"],
            "page_index": row["page_index"],
            "campo": row["campo"],
            "es_objetivo": row["es_objetivo"],
            "gold": row["gold"],
        }
        for v in ("A", "B1", "B2"):
            pred_k = f"pred_{v}"
            ok_k = f"ok_{v}"
            d_k = "—" if v == "A" else row.get(f"delta_{v}_vs_A", "")
            rows_long.append(
                {
                    **base,
                    "variante": v,
                    "pred": row.get(pred_k, ""),
                    "ok": row.get(ok_k, ""),
                    "delta_vs_A": d_k,
                }
            )
    csv_long = _OUT / "paso3_ab_por_variante.csv"
    lk = list(rows_long[0].keys()) if rows_long else []
    with open(csv_long, "w", encoding="utf-8", newline="") as f:
        w = csv.DictWriter(f, fieldnames=lk)
        w.writeheader()
        w.writerows(rows_long)

    # Excel
    try:
        from openpyxl import Workbook
    except ImportError as e:
        raise SystemExit("pip install openpyxl") from e

    xlsx_path = _OUT / "paso3_ab_revision.xlsx"
    wb = Workbook()
    ws0 = wb.active
    ws0.title = "RESUMEN"
    ws0.append(
        [
            "criterio",
            "A_target_correct",
            "B1_net_delta_target",
            "B1_regressions_non_target",
            "B2_net_delta_target",
            "B2_regressions_non_target",
            "decision_paso3_sigue",
            "variante_ganadora_regla",
        ]
    )
    ws0.append(
        [
            ">=3 target net y <=1 reg NT",
            a_tgt,
            b1_net,
            b1_reg,
            b2_net,
            b2_reg,
            "Sí" if paso3_sigue else "No (foco PASO 4)",
            mejor,
        ]
    )

    ws1 = wb.create_sheet("DETALLE_CAMPOS")
    if rows_detail:
        h = list(rows_detail[0].keys())
        ws1.append(h)
        for row in rows_detail:
            ws1.append([row.get(k, "") for k in h])

    ws2 = wb.create_sheet("POR_VARIANTE")
    if rows_long:
        h2 = list(rows_long[0].keys())
        ws2.append(h2)
        for row in rows_long:
            ws2.append([row.get(k, "") for k in h2])

    wb.save(xlsx_path)

    decision_md = (
        "## Decisión final\n\n"
        f"- **Aciertos campos objetivo (A):** {a_tgt}\n"
        f"- **B1:** delta neto objetivo vs A = **{b1_net}**, regresiones no objetivo = **{b1_reg}**\n"
        f"- **B2:** delta neto objetivo vs A = **{b2_net}**, regresiones no objetivo = **{b2_reg}**\n\n"
    )
    if paso3_sigue:
        decision_md += (
            "**PASO 3 sí vale la pena seguir** (al menos una variante cumple +3 neto en objetivos "
            f"y ≤1 regresión no objetivo). Variante bajo regla: **{mejor}**.\n"
        )
    else:
        decision_md += (
            "**PASO 3 no justifica más iteración** con esta ronda: el foco debe pasar a **PASO 4** "
            "(parsing/heurísticas). Revisar columna `paso4_candidato_A` en el CSV/Excel.\n"
        )
    if any(
        r.get("paso4_candidato_A") == "Sí" and r.get("campo") in TARGET_FIELDS for r in rows_detail
    ):
        decision_md += (
            "\n**Evidencia PASO 4:** Algún campo objetivo tiene `paso4_candidato_A=Sí` "
            "(valor gold presente en el OCR de A; fallo del extractor mínimo, no de imagen).\n"
        )

    md_path = _OUT / "INFORME_PASO3_AB_MINI.md"
    md_path.write_text(
        "\n".join(
            [
                f"# PASO 3 — Mini A/B (piloto, {run_date})",
                "",
                f"Generado: {ts}",
                "",
                "## Alcance",
                "",
                "- Páginas: `sol-viat-debedsar-2026.pdf` p.2; `rend-debedsar-amiquero-2026.pdf` p.5 y p.19.",
                "- Motor: **tesseract** (`pytesseract`, spa+eng) sobre raster **300 dpi**.",
                "- **A:** gris sin CLAHE (control sin preprocesado adicional).",
                f"- **B1:** gris + CLAHE (clipLimit={CLAHE_CLIP}, tileGridSize={CLAHE_TILE}).",
                f"- **B2:** B1 + unsharp (sigma={UNSHARP_SIGMA}, amount={UNSHARP_AMOUNT}).",
                "",
                "## Regla aplicada",
                "",
                "- Éxito: B1 o B2 con **≥ +3** aciertos netos en campos objetivo vs A, y **≤ 1** regresión en campos no objetivo.",
                "- Campos objetivo: `serie_numero`, `monto_subtotal`, `monto_total`, `monto_igv`.",
                "",
                decision_md,
                "",
                "## Artefactos",
                "",
                f"- Rasters: `{(_OUT / 'rasters').relative_to(_REPO)}`",
                f"- OCR txt: `{(_OUT / 'ocr_txt').relative_to(_REPO)}`",
                f"- CSV ancho (delta por campo): `{csv_path.relative_to(_REPO)}`",
                f"- CSV largo (variante): `{csv_long.relative_to(_REPO)}`",
                f"- Excel: `{xlsx_path.relative_to(_REPO)}`",
                "",
            ]
        ),
        encoding="utf-8",
    )

    print("OK", _OUT)
    print("CSV:", csv_path)
    print("CSV long:", csv_long)
    print("XLSX:", xlsx_path)
    print("MD:", md_path)
    print("paso3_sigue:", paso3_sigue)


if __name__ == "__main__":
    main()
