"""
PASO 2 — Bake-off con salida humana: Excel, CSV consolidado, informe MD, textos por motor.

Motores: tesseract_baseline, docling (CPU), paddleocr (si el entorno lo permite).

Métricas alineadas a docs/ROADMAP_PROYECTO.md §4.3 por campo (P, R, F1 con FP/FN/TP),
más segundos/página y % páginas con al menos un fallo (gold no null mal predicho).

Uso (Python 3.12 recomendado):
  pip install -r scripts/requirements-ocr.txt -r scripts/requirements-paso2.txt openpyxl
  export PYTHONPATH=scripts
  python scripts/bakeoff_paso2_human_export.py
  python scripts/bakeoff_paso2_human_export.py --tag linux_wsl --entorno "WSL2 Ubuntu (preferente D-12)"
"""

from __future__ import annotations

import argparse
import csv
import json
import shutil
import sys
import tempfile
import time
from collections import defaultdict
from datetime import date, datetime, timezone
from pathlib import Path
from typing import Any

_REPO = Path(__file__).resolve().parent.parent
_METRICS = _REPO / "data" / "piloto_ocr" / "metrics"
_RAW = _REPO / "data" / "piloto_ocr" / "raw"
_LABELS = _REPO / "data" / "piloto_ocr" / "labels"
_MANIFEST = _REPO / "data" / "piloto_ocr" / "MANIFEST_PILOTO.csv"

_FIELD_KEYS = [
    "ruc_emisor",
    "tipo_documento",
    "serie_numero",
    "fecha_emision",
    "moneda",
    "monto_subtotal",
    "monto_igv",
    "monto_total",
    "ruc_receptor",
    "razon_social_emisor",
]


def _load_pilot_pages() -> list[tuple[str, Path, str, int]]:
    """(doc_id, pdf_path, nombre_pdf, page_1based)."""
    out: list[tuple[str, Path, str, int]] = []
    with open(_MANIFEST, encoding="utf-8", newline="") as f:
        reader = csv.DictReader(f)
        for row in reader:
            doc_id = row["doc_id"].strip()
            rel = row["archivo_raw_relativo"].strip().replace("\\", "/")
            name = rel.split("/")[-1]
            pdf_path = _RAW / name
            if not pdf_path.is_file():
                raise FileNotFoundError(pdf_path)
            raw_indices = row["indices_paginas_en_pdf"].strip().strip('"')
            for part in raw_indices.split(","):
                part = part.strip()
                if not part:
                    continue
                out.append((doc_id, pdf_path, name, int(part)))
    if len(out) != 15:
        raise RuntimeError(f"Se esperaban 15 páginas piloto, hay {len(out)}")
    return out


def _single_page_pdf(src: Path, page_1based: int, out_path: Path) -> None:
    import fitz

    doc = fitz.open(str(src))
    new_doc = fitz.open()
    new_doc.insert_pdf(doc, from_page=page_1based - 1, to_page=page_1based - 1)
    new_doc.save(str(out_path))
    new_doc.close()
    doc.close()


def _page_bgr_300dpi(pdf_path: Path, page_1based: int):
    import cv2
    import fitz
    import numpy as np

    doc = fitz.open(str(pdf_path))
    page = doc.load_page(page_1based - 1)
    pix = page.get_pixmap(matrix=fitz.Matrix(300 / 72, 300 / 72), alpha=False)
    arr = np.frombuffer(pix.samples, dtype=np.uint8).reshape(pix.height, pix.width, pix.n)
    if pix.n == 3:
        arr = cv2.cvtColor(arr, cv2.COLOR_RGB2BGR)
    elif pix.n == 4:
        arr = cv2.cvtColor(arr, cv2.COLOR_RGBA2BGR)
    doc.close()
    return arr


def _text_baseline(pdf_one: Path) -> str:
    if str(_REPO / "scripts") not in sys.path:
        sys.path.insert(0, str(_REPO / "scripts"))
    from document_ocr_runner import process_pdf

    rows = process_pdf(str(pdf_one))
    if not rows:
        return ""
    return (rows[0].get("texto") or "").strip()


def _docling_converter():
    from docling.document_converter import DocumentConverter, PdfFormatOption
    from docling.datamodel.accelerator_options import AcceleratorDevice, AcceleratorOptions
    from docling.datamodel.base_models import InputFormat
    from docling.datamodel.pipeline_options import PdfPipelineOptions

    opts = PdfPipelineOptions()
    opts.accelerator_options = AcceleratorOptions(device=AcceleratorDevice.CPU, num_threads=4)
    return DocumentConverter(
        format_options={InputFormat.PDF: PdfFormatOption(pipeline_options=opts)}
    )


def _text_docling(conv: Any, pdf_one: Path) -> str:
    res = conv.convert(str(pdf_one))
    return (res.document.export_to_markdown() or "").strip()


def _flatten_paddle_text(result: Any) -> str:
    """Extrae líneas del resultado de PaddleOCR 3.x (estructura anidada)."""
    lines: list[str] = []

    def walk(x: Any) -> None:
        if x is None:
            return
        if isinstance(x, str) and x.strip():
            lines.append(x.strip())
            return
        if isinstance(x, dict):
            for k in ("rec_texts", "text", "transcription"):
                v = x.get(k)
                if isinstance(v, list):
                    for t in v:
                        walk(t)
                elif isinstance(v, str):
                    walk(v)
            for v in x.values():
                if isinstance(v, (list, dict, tuple)):
                    walk(v)
            return
        if isinstance(x, (list, tuple)):
            for it in x:
                walk(it)

    walk(result)
    return "\n".join(lines).strip()


def _text_paddle_image(ocr: Any, img_bgr: Any) -> tuple[str | None, str | None]:
    """Devuelve (texto, error)."""
    try:
        r = ocr.predict(img_bgr)
        if not r:
            return "", None
        return _flatten_paddle_text(r), None
    except Exception as exc:
        return None, str(exc)


def _load_gold(doc_id: str, page_index: int) -> dict[str, Any]:
    p = _LABELS / f"{doc_id}_p{page_index}.json"
    with open(p, encoding="utf-8") as f:
        return json.load(f)


def compute_field_prf(
    campo: str,
    pairs: list[tuple[Any, Any]],
) -> tuple[float, float, float, int, int, int]:
    sys.path.insert(0, str(_REPO / "scripts"))
    from piloto_field_extract_minimal import gold_pred_match

    tp = fp = fn = 0
    for gold, pred in pairs:
        g_null = gold is None or (isinstance(gold, str) and str(gold).strip() == "")
        p_null = pred is None or (isinstance(pred, str) and str(pred).strip() == "")

        if g_null:
            if not p_null:
                fp += 1
            continue
        if gold_pred_match(campo, gold, pred):
            tp += 1
        else:
            fn += 1

    prec = tp / (tp + fp) if (tp + fp) > 0 else (1.0 if fp == 0 else 0.0)
    rec = tp / (tp + fn) if (tp + fn) > 0 else 0.0
    if prec + rec == 0:
        f1 = 0.0
    else:
        f1 = 2 * prec * rec / (prec + rec)
    return prec, rec, f1, tp, fp, fn


def main(
    *,
    env_tag: str | None = None,
    entorno_label: str = "local (sin etiqueta)",
) -> None:
    if hasattr(sys.stdout, "reconfigure"):
        try:
            sys.stdout.reconfigure(encoding="utf-8", errors="replace")
        except Exception:
            pass

    sys.path.insert(0, str(_REPO / "scripts"))
    from piloto_field_extract_minimal import extract_fields_minimal

    run_date = date.today().isoformat()
    ymd = run_date.replace("-", "")
    ts = datetime.now(timezone.utc).astimezone().strftime("%Y-%m-%d %H:%M:%S %z")
    file_suffix = f"_{env_tag}" if env_tag else ""
    out_paso2 = _METRICS / (f"paso2_{env_tag}" if env_tag else "paso2")

    pages = _load_pilot_pages()
    out_paso2.mkdir(parents=True, exist_ok=True)
    for m in ("tesseract_baseline", "docling", "paddleocr"):
        (out_paso2 / m).mkdir(parents=True, exist_ok=True)

    tmpdir = tempfile.mkdtemp(prefix="paso2hv_", dir=str(_METRICS))

    docling_conv = None
    paddle_ocr: Any = None
    paddle_available = False
    paddle_fail_reason: str | None = None

    per_motor_pages: dict[str, list[dict[str, Any]]] = defaultdict(list)
    motor_times: dict[str, list[float]] = defaultdict(list)

    try:
        try:
            from paddleocr import PaddleOCR

            paddle_ocr = PaddleOCR(lang="es")
            doc_id0, pdf0, _, p0 = pages[0]
            img0 = _page_bgr_300dpi(pdf0, p0)
            _, err = _text_paddle_image(paddle_ocr, img0)
            if err:
                raise RuntimeError(err)
            paddle_available = True
        except Exception as exc:
            paddle_available = False
            paddle_fail_reason = str(exc)
            paddle_ocr = None

        docling_conv = _docling_converter()

        for doc_id, pdf_path, nombre_pdf, page_idx in pages:
            one_pdf = Path(tmpdir) / f"{doc_id}_p{page_idx}.pdf"
            _single_page_pdf(pdf_path, page_idx, one_pdf)

            gold_j = _load_gold(doc_id, page_idx)
            gold_campos = gold_j["campos"]
            ruta_rel = f"data/piloto_ocr/raw/{nombre_pdf}"

            # --- tesseract_baseline ---
            t0 = time.perf_counter()
            txt_t = _text_baseline(one_pdf)
            dt_t = time.perf_counter() - t0
            motor_times["tesseract_baseline"].append(dt_t)
            pred_t = extract_fields_minimal(txt_t)
            (out_paso2 / "tesseract_baseline" / f"{doc_id}_p{page_idx}.txt").write_text(
                txt_t, encoding="utf-8"
            )
            per_motor_pages["tesseract_baseline"].append(
                {
                    "doc_id": doc_id,
                    "nombre_pdf": nombre_pdf,
                    "ruta_relativa_repo": ruta_rel,
                    "page_index": page_idx,
                    "texto_len": len(txt_t),
                    "tiempo_s": round(dt_t, 4),
                    "gold": gold_campos,
                    "pred": pred_t,
                }
            )

            # --- docling ---
            t0 = time.perf_counter()
            txt_d = _text_docling(docling_conv, one_pdf)
            dt_d = time.perf_counter() - t0
            motor_times["docling"].append(dt_d)
            pred_d = extract_fields_minimal(txt_d)
            (out_paso2 / "docling" / f"{doc_id}_p{page_idx}.txt").write_text(
                txt_d, encoding="utf-8"
            )
            per_motor_pages["docling"].append(
                {
                    "doc_id": doc_id,
                    "nombre_pdf": nombre_pdf,
                    "ruta_relativa_repo": ruta_rel,
                    "page_index": page_idx,
                    "texto_len": len(txt_d),
                    "tiempo_s": round(dt_d, 4),
                    "gold": gold_campos,
                    "pred": pred_d,
                }
            )

            # --- paddleocr ---
            if paddle_ocr is not None:
                t0 = time.perf_counter()
                img = _page_bgr_300dpi(pdf_path, page_idx)
                txt_p, err_p = _text_paddle_image(paddle_ocr, img)
                dt_p = time.perf_counter() - t0
                if err_p or txt_p is None:
                    txt_p = ""
                motor_times["paddleocr"].append(dt_p)
                pred_p = extract_fields_minimal(txt_p or "")
                (out_paso2 / "paddleocr" / f"{doc_id}_p{page_idx}.txt").write_text(
                    (txt_p or "") + (f"\n\n[warn predict: {err_p}]\n" if err_p else ""),
                    encoding="utf-8",
                )
                per_motor_pages["paddleocr"].append(
                    {
                        "doc_id": doc_id,
                        "nombre_pdf": nombre_pdf,
                        "ruta_relativa_repo": ruta_rel,
                        "page_index": page_idx,
                        "texto_len": len(txt_p or ""),
                        "tiempo_s": round(dt_p, 4),
                        "gold": gold_campos,
                        "pred": pred_p,
                    }
                )
            else:
                (out_paso2 / "paddleocr" / f"{doc_id}_p{page_idx}.txt").write_text(
                    f"[PaddleOCR no ejecutado: {paddle_fail_reason or 'N/A'}]\n",
                    encoding="utf-8",
                )
                per_motor_pages["paddleocr"].append(
                    {
                        "doc_id": doc_id,
                        "nombre_pdf": nombre_pdf,
                        "ruta_relativa_repo": ruta_rel,
                        "page_index": page_idx,
                        "texto_len": 0,
                        "tiempo_s": 0.0,
                        "gold": gold_campos,
                        "pred": {k: None for k in _FIELD_KEYS},
                        "omitido": True,
                    }
                )

    finally:
        shutil.rmtree(tmpdir, ignore_errors=True)

    from piloto_field_extract_minimal import gold_pred_match

    def page_has_failure(gold: dict, pred: dict) -> bool:
        for campo in _FIELD_KEYS:
            g = gold.get(campo)
            p = pred.get(campo)
            if g is None or (isinstance(g, str) and not str(g).strip()):
                continue
            if not gold_pred_match(campo, g, p):
                return True
        return False

    # Métricas por motor y campo
    resumen_motor: list[dict[str, Any]] = []
    detalle_campo_rows: list[dict[str, Any]] = []
    consolidated_rows: list[dict[str, Any]] = []

    for motor, plist in per_motor_pages.items():
        if not plist:
            continue
        # % páginas fallo
        fails = 0
        for rec in plist:
            if rec.get("omitido"):
                fails += 1
                continue
            if page_has_failure(rec["gold"], rec["pred"]):
                fails += 1
        pct_fallo = 100.0 * fails / 15.0

        t_total = sum(motor_times.get(motor, [0.0]))
        sec_pp = t_total / 15.0 if motor_times.get(motor) else 0.0

        f1s: list[float] = []
        precs: list[float] = []
        recs: list[float] = []

        for campo in _FIELD_KEYS:
            if all(rec.get("omitido") for rec in plist):
                prec = rec = f1 = 0.0
                tp = fp = fn = 0
            else:
                pairs_valid = [
                    (rec["gold"].get(campo), rec["pred"].get(campo))
                    for rec in plist
                    if not rec.get("omitido")
                ]
                prec, rec, f1, tp, fp, fn = compute_field_prf(campo, pairs_valid)
            f1s.append(f1)
            precs.append(prec)
            recs.append(rec)
            detalle_campo_rows.append(
                {
                    "motor": motor,
                    "campo": campo,
                    "precision": round(prec, 4),
                    "recall": round(rec, 4),
                    "f1": round(f1, 4),
                    "tp": tp,
                    "fp": fp,
                    "fn": fn,
                }
            )

        macro_f1 = sum(f1s) / len(f1s) if f1s else 0.0
        macro_p = sum(precs) / len(precs) if precs else 0.0
        macro_r = sum(recs) / len(recs) if recs else 0.0

        disp = True
        if motor == "paddleocr":
            disp = paddle_ocr is not None
        resumen_motor.append(
            {
                "entorno_ejecucion": entorno_label,
                "motor": motor,
                "disponible": disp,
                "precision_macro_media_campo": round(macro_p, 4),
                "recall_macro_media_campo": round(macro_r, 4),
                "macro_f1_media_campo": round(macro_f1, 4),
                "segundos_por_pagina": round(sec_pp, 4),
                "tiempo_total_s": round(t_total, 2),
                "pct_paginas_con_fallo": round(pct_fallo, 2),
                "notas": (paddle_fail_reason or "")[:500] if motor == "paddleocr" and paddle_ocr is None else "",
            }
        )

    # DETALLE por página: filas largas + consolidado largo
    detalle_pagina: list[dict[str, Any]] = []
    for motor, plist in per_motor_pages.items():
        for rec in plist:
            g = rec["gold"]
            p = rec["pred"]
            obs = []
            if rec.get("omitido"):
                obs.append("PaddleOCR no ejecutado en este entorno.")
            fallo = page_has_failure(g, p) if not rec.get("omitido") else True
            row = {
                "entorno_ejecucion": entorno_label,
                "motor": motor,
                "doc_id": rec["doc_id"],
                "nombre_pdf": rec["nombre_pdf"],
                "ruta_relativa_repo": rec["ruta_relativa_repo"],
                "ruta_absoluta_pdf": str((_REPO / rec["ruta_relativa_repo"]).resolve()),
                "page_index": rec["page_index"],
                "tiempo_pagina_s": rec["tiempo_s"],
                "len_texto": rec["texto_len"],
                "pagina_con_fallo": "Sí" if fallo else "No",
                "observaciones": "; ".join(obs) if obs else "",
            }
            for campo in _FIELD_KEYS:
                gv = g.get(campo)
                pv = p.get(campo) if p else None
                ok = ""
                if gv is None or (isinstance(gv, str) and not str(gv).strip()):
                    ok = "N/A"
                elif rec.get("omitido"):
                    ok = "ERR"
                else:
                    ok = "1" if gold_pred_match(campo, gv, pv) else "0"
                row[f"gold_{campo}"] = gv if gv is not None else ""
                row[f"pred_{campo}"] = pv if pv is not None else ""
                row[f"ok_{campo}"] = ok
            detalle_pagina.append(row)

            for campo in _FIELD_KEYS:
                gv = g.get(campo)
                pv = p.get(campo) if p else None
                ex = ""
                if gv is None or (isinstance(gv, str) and not str(gv).strip()):
                    ex = "N/A"
                elif rec.get("omitido"):
                    ex = "0"
                else:
                    ex = "1" if gold_pred_match(campo, gv, pv) else "0"
                consolidated_rows.append(
                    {
                        "entorno_ejecucion": entorno_label,
                        "motor": motor,
                        "doc_id": rec["doc_id"],
                        "nombre_pdf": rec["nombre_pdf"],
                        "ruta_relativa_repo": rec["ruta_relativa_repo"],
                        "ruta_absoluta_pdf": str((_REPO / rec["ruta_relativa_repo"]).resolve()),
                        "page_index": rec["page_index"],
                        "campo": campo,
                        "valor_gold": gv if gv is not None else "",
                        "valor_pred": pv if pv is not None else "",
                        "exactitud_binaria": ex,
                        "tiempo_pagina_s": rec["tiempo_s"],
                        "len_texto_pagina": rec["texto_len"],
                        "fecha_corrida": run_date,
                    }
                )

    # Guardar CSV consolidado
    csv_path = _METRICS / f"bakeoff_paso2_consolidado_{ymd}{file_suffix}.csv"
    if consolidated_rows:
        keys = list(consolidated_rows[0].keys())
        with open(csv_path, "w", encoding="utf-8", newline="") as f:
            w = csv.DictWriter(f, fieldnames=keys)
            w.writeheader()
            w.writerows(consolidated_rows)

    # Excel
    try:
        from openpyxl import Workbook
    except ImportError as e:
        raise SystemExit("Instale openpyxl: pip install openpyxl") from e

    xlsx_path = _METRICS / f"bakeoff_paso2_revision_{ymd}{file_suffix}.xlsx"
    wb = Workbook()
    # NOTA_ENTORNO
    ws_note = wb.active
    ws_note.title = "NOTA_ENTORNO"
    ws_note.append(["entorno_ejecucion", entorno_label])
    ws_note.append(["env_tag", env_tag or "(default paso2/)"])
    ws_note.append(["textos_intermedios", str(out_paso2.relative_to(_REPO))])
    ws_note.append(
        [
            "comparacion_preferente",
            "Si env_tag=linux_wsl: esta corrida es la comparacion preferente (D-12). "
            "Corrida Windows referencia: bakeoff_paso2_revision_20260413.xlsx (mismo piloto).",
        ]
    )
    # RESUMEN
    ws0 = wb.create_sheet("RESUMEN", 1)
    h0 = list(resumen_motor[0].keys()) if resumen_motor else []
    if h0:
        ws0.append(h0)
        for row in resumen_motor:
            ws0.append([row.get(k, "") for k in h0])
    # RESUMEN_CAMPOS
    ws1 = wb.create_sheet("RESUMEN_CAMPOS")
    if detalle_campo_rows:
        h1 = list(detalle_campo_rows[0].keys())
        ws1.append(h1)
        for row in detalle_campo_rows:
            ws1.append([row.get(k, "") for k in h1])
    # DETALLE_POR_PAGINA
    ws2 = wb.create_sheet("DETALLE_POR_PAGINA")
    if detalle_pagina:
        h2 = list(detalle_pagina[0].keys())
        ws2.append(h2)
        for row in detalle_pagina:
            ws2.append([row.get(k, "") for k in h2])
    # Una hoja por motor
    for motor in ("tesseract_baseline", "docling", "paddleocr"):
        rows_m = [r for r in detalle_pagina if r["motor"] == motor]
        wsm = wb.create_sheet(motor[:31])
        if rows_m:
            hm = list(rows_m[0].keys())
            wsm.append(hm)
            for row in rows_m:
                wsm.append([row.get(k, "") for k in hm])
    wb.save(xlsx_path)

    # Markdown
    md_path = _METRICS / f"INFORME_PASO2_REVISION_{ymd}{file_suffix}.md"
    lines = [
        f"# Informe PASO 2 — Bake-off revisable ({run_date})",
        "",
        f"**Entorno de ejecución:** {entorno_label}",
        "",
        f"**Etiqueta de corrida (`--tag`):** `{env_tag or '—'}` — CSV/XLSX/MD con sufijo `{file_suffix or '(ninguno)'}`. Textos por motor: `{out_paso2.relative_to(_REPO)}/<motor>/`.",
        "",
        f"Generado: {ts}",
        "",
    ]
    if env_tag == "linux_wsl":
        lines.extend(
            [
                "## Comparación preferente (D-12)",
                "",
                "Esta corrida en **Linux/Ubuntu vía WSL** se trata como la **comparación preferente y más confiable** frente a la corrida previa en **Windows** (artefactos `bakeoff_paso2_*_20260413.*` / `INFORME_PASO2_REVISION_20260413.md` sin sufijo `linux_wsl`), cuando Windows pudo introducir limitaciones de runtime (p. ej. PaddleOCR / oneDNN).",
                "",
                "### Referencia Windows (2026-04-13)",
                "",
                "| motor | P macro | R macro | F1 macro | s/página | % págs. fallo |",
                "|-------|---------|---------|----------|----------|----------------|",
                "| tesseract_baseline | 0.9833 | 0.6572 | 0.7388 | 0.0096 | 100.0 |",
                "| docling | 0.9833 | 0.5761 | 0.6852 | 1.5756 | 100.0 |",
                "| paddleocr | 0.0 | 0.0 | 0.0 | 0.0 | 100.0 |",
                "",
                "*Valores tomados del informe Windows `INFORME_PASO2_REVISION_20260413.md` (Paddle no ejecutó inferencia en ese entorno).*",
                "",
            ]
        )
    lines.extend(
        [
            "## Motores",
            "",
            "- **tesseract_baseline**: ejecutado.",
            "- **docling (CPU)**: ejecutado.",
            f"- **paddleocr**: {'ejecutado' if paddle_ocr is not None else 'no viable / error en este entorno'}",
            "",
        ]
    )
    if paddle_fail_reason:
        lines.extend(
            [
                "### PaddleOCR",
                "",
                f"```\n{paddle_fail_reason[:2000]}\n```",
                "",
            ]
        )
    if env_tag == "linux_wsl":
        lines.extend(
            [
                "## Esta corrida (WSL/Linux)",
                "",
                "### Tabla comparativa (resumen)",
                "",
            ]
        )
    else:
        lines.extend(
            [
                "## Tabla comparativa (resumen)",
                "",
            ]
        )
    lines.extend(
        [
            "| motor | P macro* | R macro* | F1 macro* | s/página | % págs. fallo |",
            "|-------|----------|----------|-----------|----------|----------------|",
        ]
    )
    for row in resumen_motor:
        lines.append(
            f"| {row['motor']} | {row.get('precision_macro_media_campo', '')} | "
            f"{row.get('recall_macro_media_campo', '')} | {row.get('macro_f1_media_campo', '')} | "
            f"{row.get('segundos_por_pagina', '')} | {row.get('pct_paginas_con_fallo', '')} |"
        )
    lines.extend(
        [
            "",
            "*Media aritmética de precision/recall/F1 **por campo** (10 campos del extractor mínimo).",
            "",
        ]
    )
    if env_tag == "linux_wsl":
        _p2 = out_paso2.relative_to(_REPO)
        paddle_wsl = (
            f"PaddleOCR **sí** ejecutó inferencia en WSL (hoja Excel `paddleocr`, textos `{_p2}/paddleocr/`)."
            if paddle_ocr is not None
            else f"PaddleOCR **no** ejecutó inferencia en WSL; ver error arriba y placeholders en `{_p2}/paddleocr/`."
        )
        lines.extend(
            [
                "## Windows vs WSL/Linux",
                "",
                "| Aspecto | Windows (corrida ref. 2026-04-13) | WSL/Linux (esta corrida) |",
                "|---------|-----------------------------------|---------------------------|",
                "| Artefactos | `bakeoff_paso2_*_20260413.*`, `paso2/<motor>/` | mismos nombres con sufijo `_linux_wsl` y carpeta `paso2_linux_wsl/` |",
                "| Confianza D-12 | secundaria si el runtime distorsionó OCR | **preferente** para decisión de motores |",
                "| PaddleOCR | no inferencia (0 métricas útiles) | " + paddle_wsl.replace("|", "\\|") + " |",
                "",
            ]
        )
    lines.extend(
        [
            "## Archivos",
            "",
            f"- Excel: `{xlsx_path.relative_to(_REPO)}`",
            f"- CSV consolidado: `{csv_path.relative_to(_REPO)}`",
            f"- Textos por motor: `{out_paso2.relative_to(_REPO)}/<motor>/`",
            "",
        ]
    )
    if env_tag == "linux_wsl":
        lines.extend(
            [
                "## Recomendación preliminar actualizada",
                "",
                "Comparar la tabla **Esta corrida (WSL/Linux)** con **Referencia Windows**. Priorizar F1 macro y recall por campo según ROADMAP; si PaddleOCR solo es viable en WSL, la recomendación de motor puede cambiar respecto de Windows.",
                "",
            ]
        )
    md_path.write_text("\n".join(lines), encoding="utf-8")

    print("OK")
    print("Excel:", xlsx_path)
    print("CSV:", csv_path)
    print("MD:", md_path)


if __name__ == "__main__":
    ap = argparse.ArgumentParser(description="PASO 2 bake-off export humano")
    ap.add_argument(
        "--tag",
        default=None,
        help="Sufijo en archivos y carpeta paso2_<tag> (ej. linux_wsl)",
    )
    ap.add_argument(
        "--entorno",
        default="local (sin etiqueta)",
        help='Etiqueta legible del entorno (ej. "WSL2 Ubuntu")',
    )
    args = ap.parse_args()
    main(env_tag=args.tag, entorno_label=args.entorno)
